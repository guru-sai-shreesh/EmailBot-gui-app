import smtplib
import speech_recognition as sr
import pyttsx3
from email.message import EmailMessage
import openpyxl as xl
from kivymd.app import MDApp
from kivy.core.window import Window
from kivy.lang.builder import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.uix.list import TwoLineAvatarListItem, IconLeftWidget
import copy
# for listener to work install pyaudio package

email_receivers = []
addresses = []
sub = []
body = []
new_contacts = []

wb = xl.load_workbook('contacts.xlsx')
sheet = wb['Sheet1']
contact_list = {}
x = 2
for x in range(2, sheet.max_row + 1):
    cell1 = sheet.cell(x, 2)
    cell2 = sheet.cell(x, 3)
    contact_list[cell1.value] = cell2.value

listener = sr.Recognizer()

engine = pyttsx3.init()
engine.setProperty('rate', 180)  # reduces WPM to 180
engine.setProperty('voice', 'com.apple.speech.synthesis.voice.samantha.premium')  # This voice is best suited in macos


def talk(text):
    engine.say(text)
    engine.runAndWait()


def mike_out():
    try:
        with sr.Microphone() as source:
            print('listening...')
            voice = listener.listen(source)
            info = listener.recognize_google(voice)
            print(info)
            return info.lower()
    except:
        pass


def send_email(receiver, subject, message):
    server = smtplib.SMTP('smtp.gmail.com', 587)  # to connect to server
    server.starttls()  # tells that it is an secure connection
    # Make sure to give app access in your Google account
    server.login('gojo.testing123@gmail.com', 'hellogojo')
    email = EmailMessage()
    email['From'] = 'gojo.testing123@gmail.com'
    email['To'] = receiver
    email['Subject'] = subject
    email.set_content(message)
    server.send_message(email)


def gather_and_send():
    subject0 = ""
    body0 = ""
    i = 0
    j = 0
    for subs in sub:
        if i == 0:
            subject0 = subs.capitalize()
            i += 1
        else:
            subject0 += ('. ' + subs.capitalize())
    for bodies in body:
        if j == 0:
            body0 = bodies.capitalize()
            j += 1
        else:
            body0 += ('. ' + bodies.capitalize())
    unique_addresses = list(set(addresses))
    for receiver in unique_addresses:
        send_email(receiver, subject0, body0)
    print('Email sent successfully to', *unique_addresses)


def new_contact():
    for name, email in new_contacts:
        contact_list[name] = email
        new_cell0 = sheet.cell(x + 1, 1)
        new_cell1 = sheet.cell(x + 1, 2)
        new_cell2 = sheet.cell(x + 1, 3)
        new_cell0.value = x - 1
        new_cell1.value = name
        new_cell2.value = email
        wb.save('contacts.xlsx')
    wb.close()


screen_helper = """
ScreenManager:
    MenuScreen:
    SelectScreen:
    NCScreen:
    SubjectScreen:
    BodyScreen:
    EndScreen:
<MenuScreen>:
    name: 'menu'
    Image:
        source: 'bot_anim.gif'
        anim_delay: 0.03
        mipmap: True
        allow_stretch: True
        pos_hint: {'center_x':0.5,'center_y':0.5}
    MDLabel:
        id: head
        text: 'EMAIL BOT'
        bold: True
        halign: 'center'
        bold: True
        font_size: '30sp'
        pos_hint: {'center_x':0.34,'center_y':0.83}
        color: 0.16, 0.47, 0.45, 1
    MDRaisedButton:
        text: 'Start'
        pos_hint: {'center_x':0.8,'center_y':0.14}
        on_press: root.manager.current = 'select'
        elevation: 10
<SelectScreen>:
    name: 'select'
    BoxLayout:
        ScrollView:
            MDList:
                id: scroll
    MDRaisedButton:
        text: 'New contact'
        pos_hint: {'center_x':0.2,'center_y':0.05}
        on_press: root.manager.current = 'new_contact'
    MDRaisedButton:
        text: 'Speak names to add'
        pos_hint: {'center_x':0.58,'center_y':0.05}
        md_bg_color: app.theme_cls.primary_dark
        elevation: 12
        on_press: root.receiver_addresses()
    MDRectangleFlatButton:
        text: 'Next'
        pos_hint: {'center_x':0.88,'center_y':0.05}
        on_press: root.manager.current = 'subject'

<NCScreen>:
    name: 'new_contact'
    MDLabel:
        id: head
        text: 'NEW CONTACT'
        bold: True
        halign: 'center'
        bold: True
        font_size: '20sp'
        pos_hint: {'center_y':0.7}
        color: 0, 0, 1, 1
    MDLabel:
        id: head
        text: 'You can add multiple contacts by saving them in pairs'
        bold: True
        halign: 'center'
        bold: True
        pos_hint: {'center_y':0.6}
        color: 0.16, 0.47, 0.45, 1
    MDTextField:
        id: name
        hint_text: "Enter name"
        helper_text: "and Enter Email address"
        helper_text_mode: "on_focus"
        icon_right_color: app.theme_cls.primary_color
        pos_hint:{'center_x': 0.5, 'center_y': 0.5}
        color_mode: 'custom'
        mode: "rectangle"
        line_color_focus: 0, 0, 0.9, 1
        size_hint_x:None
        width:500
    MDTextField:
        id: address
        hint_text: "Enter email address"
        helper_text: "or click on Next"
        helper_text_mode: "on_focus"
        icon_right_color: app.theme_cls.primary_color
        pos_hint:{'center_x': 0.5, 'center_y': 0.4}
        color_mode: 'custom'
        mode: "rectangle"
        line_color_focus: 0, 0, 0.9, 1
        size_hint_x:None
        width:500
    MDRaisedButton:
        text: 'Save'
        pos_hint: {'center_x': 0.5, 'center_y': 0.3}
        on_release: root.save_data()
    MDRectangleFlatButton:
        text: 'Next'
        pos_hint: {'center_x': 0.88, 'center_y': 0.05}
        on_release: root.manager.current = 'subject'
<SubjectScreen>:
    name: 'subject'
    MDLabel:
        text: 'Subject:'
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.13,'center_y':0.95}
        color: 0.16, 0.47, 0.45, 1
    MDRectangleFlatButton:
        id: sub
        text: "Press on 'start listening' to speak subject"
        pos_hint: {'center_x':0.5,'center_y':0.5}
        size_hint: 0.9, 0.8
    MDRaisedButton:
        text: 'Start listening'
        pos_hint: {'center_x':0.8,'center_y':0.95}
        md_bg_color: app.theme_cls.primary_dark
        on_press: root.listen_subject()
    MDRaisedButton:
        text: 'Next'
        pos_hint: {'center_x':0.87,'center_y':0.05}
        on_press: root.manager.current = 'body'
<BodyScreen>:
    name: 'body'
    MDLabel:
        text: 'Email Body: '
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.18,'center_y':0.95}
        color: 0.16, 0.47, 0.45, 1
    MDRectangleFlatButton:
        id: ebod
        text: "Press on 'start listening' to speak BODY"
        pos_hint: {'center_x':0.5,'center_y':0.5}
        size_hint: 0.9, 0.8
    MDRaisedButton:
        text: 'Start listening'
        pos_hint: {'center_x':0.8,'center_y':0.95}
        md_bg_color: app.theme_cls.primary_dark
        on_press: root.listen_body()
    MDRaisedButton:
        text: 'Next'
        pos_hint: {'center_x':0.87,'center_y':0.05}
        on_press: root.manager.current = 'end'
<EndScreen>:
    name: 'end'
    MDLabel:
        text: "Press on 'Send' button to send composed Email"
        pos_hint: {'center_x':0.54,'center_y':0.66}
        color: 0.16, 0.47, 0.45, 1
    MDRaisedButton:
        text: 'Send'
        pos_hint: {'center_x':0.5,'center_y':0.6}
        on_press: root.final_send()
    MDLabel:
        id: final
        halign: 'center'
        pos_hint: {'center_x':0.5,'center_y':0.5}
    MDLabel:
        id: notify
        pos_hint: {'center_x':0.57,'center_y':0.05}
        color: 0.16, 0.47, 0.45, 1

"""


class MenuScreen(Screen):
    pass


class SelectScreen(Screen):

    def receiver_addresses(self):
        while(1):
            receivers_str = mike_out()
            receivers = receivers_str.split(' and ')
            for receiver in receivers:
                if receiver not in contact_list:
                    receivers.remove(receiver)
            if len(receivers) == 0:
                talk("sorry, there are no similar email addresses in your contacts")
                talk("Please try again")
            else:
                break

        for receiver in receivers:
            email_receivers.append(receiver.capitalize())
            addresses.append(contact_list[receiver])
        print("Receiver's Email addresses: ", *addresses)


class NCScreen(Screen):

    def save_data(self):
        new_contacts.append([self.ids.name.text.lower(), self.ids.address.text.lower()])
        addresses.append(self.ids.address.text.lower().capitalize())
        email_receivers.append([self.ids.name.text.lower().capitalize()])
        new_contact()


class SubjectScreen(Screen):
    line_count = 0

    def listen_subject(self):
        sub_line = mike_out()
        sub.append(sub_line)
        dup = copy.deepcopy(sub_line)
        c = 0
        for index in range(len(dup) - 1):
            if dup[index] == " " and c < 7:
                c += 1
            elif dup[index] == " " and c >= 7:
                dup = dup[:index] + '\n' + dup[index + 1:]
                c = 0
        if self.line_count == 0:
            self.ids.sub.text = dup.capitalize()
            self.line_count += 1
        else:
            self.ids.sub.text += f"\n{dup.capitalize()}"


class BodyScreen(Screen):
    line_count = 0

    def listen_body(self):
        body_line = mike_out()
        body.append(body_line)
        dup = copy.deepcopy(body_line)
        c = 0
        for index in range(len(dup) - 1):
            if dup[index] == " " and c < 7:
                c += 1
            elif dup[index] == " " and c >= 7:
                dup = dup[:index] + '\n' + dup[index + 1:]
                c = 0
        if self.line_count == 0:
            self.ids.ebod.text = dup.capitalize()
            self.line_count += 1
        else:
            self.ids.ebod.text += f"\n{dup.capitalize()}"


class EndScreen(Screen):

    def final_send(self):
        gather_and_send()
        final_msg_names = ""
        unique_receivers = list(set(email_receivers))
        for receiver in unique_receivers:
            final_msg_names += (' ' + receiver)
        self.ids.final.text = "Email(s) successfully sent to\n" + final_msg_names


# Create the screen manager
sm = ScreenManager()
sm.add_widget(MenuScreen(name='menu'))
sm.add_widget(SelectScreen(name='select'))
sm.add_widget(NCScreen(name='new_contact'))
sm.add_widget(SubjectScreen(name='subject'))
sm.add_widget(BodyScreen(name='body'))
sm.add_widget(EndScreen(name='end'))

Window.size = (360, 600)


class DemoApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Blue"
        screen = Screen()
        self.help_str = Builder.load_string(screen_helper)
        screen.add_widget(self.help_str)
        for key in contact_list:
            icons = IconLeftWidget(icon="contact_icon.jpeg")
            items = TwoLineAvatarListItem(text=key.capitalize(), secondary_text=contact_list[key])
            items.add_widget(icons)
            self.help_str.get_screen('select').ids.scroll.add_widget(items)
        return screen


DemoApp().run()
